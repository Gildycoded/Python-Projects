#! python3
# sendDuesReminders.py

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

#Check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.max_row +1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment !='paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
#Log into email
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
superAddress=input('Enter your email here:')
superPassword=input('Enter your password here:')
smtpObj.login(superAddress, superPassword)

#Send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid. \n Dear %s, \nRecords show you haven't paid me for %s. Fix it ASAP!'" %(latestMonth, name, latestMonth)
    print('Sending email to %s...' %email)
    sendmailStatus = smtpObj.sendmail(superAddress, email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' %(email, sendmailStatus))
smtpObj.quit()

